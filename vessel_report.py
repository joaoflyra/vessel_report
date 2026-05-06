import anthropic
import imaplib
import smtplib
import email
import email.header
import os
import io
from email.mime.text import MIMEText
from email.mime.multipart import MIMEMultipart
from datetime import datetime, timedelta

import openpyxl

GMAIL_USER = os.environ.get("GMAIL_USER", "joao@lyrashipping.com.br")
GMAIL_APP_PASSWORD = os.environ.get("GMAIL_APP_PASSWORD", "")
ANTHROPIC_API_KEY = os.environ.get("ANTHROPIC_API_KEY", "")
REPORT_RECIPIENT = os.environ.get("REPORT_RECIPIENT", GMAIL_USER)
CRISTIANO_EMAIL = "cristiano@lyrashipping.com.br"
JEROME_EMAIL = "jerome@lyrashipping.com.br"
POSITION_LIST_SENDERS = [CRISTIANO_EMAIL, JEROME_EMAIL]

FLEET = ["EKATERINA", "CALLIO", "DAHLIA", "PARAGON", "HERAKLITOS", "MARCOS DIAS", "LEFTERIS T"]


def previous_business_day():
    today = datetime.now()
    offset = 1
    if today.weekday() == 0:
        offset = 3
    elif today.weekday() == 6:
        offset = 2
    return today - timedelta(days=offset)


def decode_header_value(value):
    parts = email.header.decode_header(value)
    decoded = []
    for part, charset in parts:
        if isinstance(part, bytes):
            decoded.append(part.decode(charset or "utf-8", errors="ignore"))
        else:
            decoded.append(part)
    return "".join(decoded)


def excel_to_text(data: bytes) -> str:
    wb = openpyxl.load_workbook(io.BytesIO(data), data_only=True)
    lines = []
    for sheet in wb.sheetnames:
        ws = wb[sheet]
        lines.append(f"[Aba: {sheet}]")
        for row in ws.iter_rows(values_only=True):
            row_values = [str(v) if v is not None else "" for v in row]
            if any(v.strip() for v in row_values):
                lines.append("\t".join(row_values))
    return "\n".join(lines)


def extract_fleet_from_excel(data: bytes) -> list:
    """Extrai nomes dos navios do Excel da position list — apenas celulas em negrito na coluna vessel."""
    try:
        wb = openpyxl.load_workbook(io.BytesIO(data), data_only=True)
        SKIP_WORDS = {"vessel", "navio", "ship", "mv", "m/v", "nome", "name", "date",
                      "porto", "port", "status", "eta", "etd", "none", "n/a", ""}
        vessels = []
        for sheet in wb.sheetnames:
            ws = wb[sheet]
            vessel_col = None
            for row in ws.iter_rows(min_row=1, max_row=5):
                for idx, cell in enumerate(row):
                    if cell.value and str(cell.value).strip().lower() in ("vessel", "navio", "ship", "mv", "m/v", "nome"):
                        vessel_col = idx + 1  # openpyxl e 1-based para iter por coluna
                        break
                if vessel_col is not None:
                    break
            if vessel_col is None:
                vessel_col = 1
            for row in ws.iter_rows(min_row=2):
                if vessel_col <= len(row):
                    cell = row[vessel_col - 1]
                    if cell.value and cell.font and cell.font.bold:
                        name = str(cell.value).strip()
                        if name.lower() not in SKIP_WORDS and not name.replace(".", "").isdigit():
                            vessels.append(name.upper())
        return list(dict.fromkeys(vessels))
    except Exception:
        return []


def fetch_cristiano_positions(mail_conn):
    pbd = previous_business_day()
    since = pbd.strftime("%d-%b-%Y")

    all_ids = []
    for sender in POSITION_LIST_SENDERS:
        _, ids = mail_conn.search(None, f'(FROM "{sender}" SINCE {since})')
        if ids[0]:
            all_ids.extend(ids[0].split())

    if not all_ids:
        return None, None, []

    for msg_id in reversed(all_ids):
        _, msg_data = mail_conn.fetch(msg_id, "(RFC822)")
        msg = email.message_from_bytes(msg_data[0][1])
        subject = decode_header_value(msg.get("Subject", ""))
        date_str = msg.get("Date", "")
        sender_addr = msg.get("From", "")

        for part in msg.walk():
            fn = part.get_filename()
            if fn:
                fn_decoded = decode_header_value(fn)
                ext = fn_decoded.lower().split(".")[-1]
                if ext in ("xlsx", "xls"):
                    payload = part.get_payload(decode=True)
                    if payload:
                        text = excel_to_text(payload)
                        fleet = extract_fleet_from_excel(payload)
                        return text, f"De: {sender_addr} | Data: {date_str} | Assunto: {subject}", fleet
    return None, None, []


def fetch_last_report(mail_conn):
    """Busca o último relatório diário enviado para contextualização."""
    since = (datetime.now() - timedelta(days=7)).strftime("%d-%b-%Y")

    sent_folders = ['"[Gmail]/Sent Mail"', '"[Gmail]/Enviados"', "Sent", "Enviados"]
    sent_folder = None
    for folder in sent_folders:
        try:
            status, _ = mail_conn.select(folder)
            if status == "OK":
                sent_folder = folder
                break
        except Exception:
            continue

    if not sent_folder:
        mail_conn.select("inbox")
        return None, None

    try:
        _, ids = mail_conn.search(None, f'(SUBJECT "Relatorio Diario" SINCE {since})')
        if not ids[0]:
            mail_conn.select("inbox")
            return None, None

        msg_id = ids[0].split()[-1]
        _, msg_data = mail_conn.fetch(msg_id, "(RFC822)")
        msg = email.message_from_bytes(msg_data[0][1])
        date_str = msg.get("Date", "")

        body = ""
        if msg.is_multipart():
            for part in msg.walk():
                if part.get_content_type() == "text/plain":
                    payload = part.get_payload(decode=True)
                    if payload:
                        body = payload.decode("utf-8", errors="ignore")
                        break
        else:
            payload = msg.get_payload(decode=True)
            if payload:
                body = payload.decode("utf-8", errors="ignore")

        mail_conn.select("inbox")
        return body[:5000], date_str
    except Exception:
        mail_conn.select("inbox")
        return None, None


def fetch_fixture_recaps(mail_conn):
    """Busca fixture recaps dos ultimos 30 dias para referencia de agentes e charterers."""
    since = (datetime.now() - timedelta(days=30)).strftime("%d-%b-%Y")
    recaps = []

    for keyword in ["recap", "fixture", "charter party"]:
        try:
            _, ids = mail_conn.search(None, f'(SUBJECT "{keyword}" SINCE {since})')
            if not ids[0]:
                continue
            for msg_id in ids[0].split()[-20:]:
                _, msg_data = mail_conn.fetch(msg_id, "(RFC822)")
                msg = email.message_from_bytes(msg_data[0][1])
                subject = decode_header_value(msg.get("Subject", ""))
                sender = decode_header_value(msg.get("From", ""))
                date_str = msg.get("Date", "")
                body = ""
                if msg.is_multipart():
                    for part in msg.walk():
                        if part.get_content_type() == "text/plain":
                            payload = part.get_payload(decode=True)
                            if payload:
                                body = payload.decode("utf-8", errors="ignore")
                                break
                else:
                    payload = msg.get_payload(decode=True)
                    if payload:
                        body = payload.decode("utf-8", errors="ignore")
                recaps.append({"subject": subject, "from": sender, "date": date_str, "body": body[:2000]})
        except Exception:
            continue

    # deduplicar por assunto+remetente
    seen = set()
    unique = []
    for r in recaps:
        key = (r["subject"], r["from"])
        if key not in seen:
            seen.add(key)
            unique.append(r)
    return unique


def fetch_vessel_emails():
    mail = imaplib.IMAP4_SSL("imap.gmail.com")
    mail.login(GMAIL_USER, GMAIL_APP_PASSWORD)
    mail.select("inbox")

    positions_text, positions_meta, fleet_from_excel = fetch_cristiano_positions(mail)
    last_report, last_report_date = fetch_last_report(mail)
    fixture_recaps = fetch_fixture_recaps(mail)

    since_date = (datetime.now() - timedelta(days=1)).strftime("%d-%b-%Y")
    _, message_ids = mail.search(None, f"(SINCE {since_date})")

    emails = []
    ids = message_ids[0].split()
    for msg_id in ids[-50:]:
        _, msg_data = mail.fetch(msg_id, "(RFC822)")
        msg = email.message_from_bytes(msg_data[0][1])

        subject = decode_header_value(msg.get("Subject", "(sem assunto)"))
        sender = decode_header_value(msg.get("From", ""))
        date_str = msg.get("Date", "")
        body = ""

        if msg.is_multipart():
            for part in msg.walk():
                if part.get_content_type() == "text/plain":
                    payload = part.get_payload(decode=True)
                    if payload:
                        body = payload.decode("utf-8", errors="ignore")
                        break
        else:
            payload = msg.get_payload(decode=True)
            if payload:
                body = payload.decode("utf-8", errors="ignore")

        body = body[:3000]
        emails.append({"subject": subject, "from": sender, "date": date_str, "body": body})

    mail.logout()
    return emails, positions_text, positions_meta, fleet_from_excel, last_report, last_report_date, fixture_recaps


def generate_report(emails, positions_text, positions_meta, fleet_from_excel, last_report, last_report_date, fixture_recaps):
    client = anthropic.Anthropic(api_key=ANTHROPIC_API_KEY)

    today = datetime.now().strftime("%d/%m/%Y")
    active_fleet = fleet_from_excel if fleet_from_excel else FLEET
    fleet_list = ", ".join(active_fleet)

    email_content = "\n\n---\n\n".join([
        f"De: {e['from']}\nData: {e['date']}\nAssunto: {e['subject']}\n\n{e['body']}"
        for e in emails
    ]) if emails else "Nenhum email recebido nas ultimas 24 horas."

    if positions_text:
        positions_section = f"""
━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━
PLANILHA DE POSICOES (referencia de nomes apenas)
{positions_meta}
━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━
USE esta planilha APENAS para identificar quais navios da frota estao ativos.
NAO extraia informacoes operacionais desta planilha (rota, ETA, status, bunkers, etc.).
Toda informacao operacional deve vir exclusivamente dos emails.
{positions_text[:4000]}
"""
    else:
        positions_section = "\nPlanilha de posicoes do Cristiano: nao encontrada.\n"

    if last_report:
        last_report_section = f"""
━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━
ULTIMO RELATORIO ENVIADO (para contextualizacao)
Data: {last_report_date}
━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━
{last_report}
"""
    else:
        last_report_section = "\nUltimo relatorio: nao encontrado.\n"

    if fixture_recaps:
        recap_content = "\n\n---\n\n".join([
            f"De: {r['from']}\nData: {r['date']}\nAssunto: {r['subject']}\n\n{r['body']}"
            for r in fixture_recaps
        ])
        recaps_section = f"""
━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━
FIXTURE RECAPS / CHARTER PARTIES (ultimos 30 dias — referencia para agentes e charterers)
━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━
Use estes documentos para confirmar agentes e charterers em caso de duvida.
{recap_content[:6000]}
"""
    else:
        recaps_section = "\nFixture recaps: nenhum encontrado nos ultimos 30 dias.\n"

    prompt = f"""Voce e um assistente especializado em operacoes maritimas da empresa Lyra Shipping.

REGRAS FUNDAMENTAIS:

1. SEM DEDUCOES: Inclua APENAS informacoes confirmadas explicitamente nos emails.
   - Jamais deduza, infira ou suponha qualquer informacao.
   - Se o ETA nao estiver no email, omita o campo ETA silenciosamente.
   - Em caso de duvida, omita ou indique "dado nao confirmado".

2. NOMES DE PORTOS: Copie o nome do porto EXATAMENTE como aparece no email.
   - Nunca abrevie, trunce ou interprete nomes de portos.
   - Se o porto nao estiver claro no email, escreva "porto nao especificado".

3. BUNKERS: Informe quantidade de bunker SOMENTE em dois momentos-chave:
   a) Abastecimento realizado — o navio acabou de abastecer (confirmar quantidade e porto nos emails).
   b) Necessidade iminente de decisao — nivel critico confirmado nos emails que exige acao (cotacao solicitada, pedido em andamento, agente alertando para nivel baixo).
   - NUNCA informe ROB como dado rotineiro de status.
   - NUNCA mencione bunkers com base na planilha de posicoes.
   - Fora desses dois momentos, omita completamente qualquer referencia a bunkers.

4. PLANILHA DE POSICOES: Use-a APENAS para identificar quais navios estao ativos.
   - NAO extraia informacoes operacionais da planilha (rota, ETA, status, bunkers).
   - Toda informacao operacional vem exclusivamente dos emails.

5. AGENTE vs FRETADOR: Nunca confundir agente maritimo com fretador/charterer.
   - Agente = empresa local que representa o navio no porto (ex: HMS Brasil, Fratino & Figli).
   - Fretador/Charterer = empresa que contratou o navio (confirmar sempre no fixture recap).
   - Se nao tiver certeza do charterer, omitir — nao deduzir pelo nome do agente.

OUTRAS REGRAS:
- NAO use markdown (sem ##, sem **, sem _).
- Use APENAS texto simples com os separadores abaixo.
- Seja conciso. Sem frases longas. Prefira bullet points curtos.
- Evite siglas tecnicas sem explicacao.
- PROXIMA VIAGEM = apenas a viagem imediatamente apos a atual. Nunca mencionar a proxima da proxima, mesmo que ja esteja fechada.
- Negociacoes em andamento confirmadas nos emails: mencionar com 🔄 Em negociacao, brevemente. Nunca tratar como fixture confirmado.
- Nao repita informacoes do ultimo relatorio se nao houver atualizacao confirmada.
- Sem duplicidades entre secoes: cada informacao aparece UMA unica vez, na secao mais adequada.
- Questoes de viagens ANTERIORES (laytime, disputas, faturas, demurrage de voyages encerradas) NUNCA entram no bloco do navio.
  Estas questoes vao EXCLUSIVAMENTE na secao de pendencias no final. Sem excecoes.
- Quando uma informacao nao estiver disponivel (contato, ETA, etc.), omitir silenciosamente — sem explicar entre parenteses o motivo da ausencia.

A frota tem EXATAMENTE estes {len(active_fleet)} navios — inclua TODOS, sem excecao:
{fleet_list}

{positions_section}

{last_report_section}

{recaps_section}

━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━
FORMATO DO RELATORIO — use EXATAMENTE:
━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━

━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━
LYRA SHIPPING  ·  RELATORIO DIARIO  ·  {today}
━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━━

[Para cada navio:]

[NIVEL]  M/V [NOME]  —  VOY [XXX/XX]
  Status    [navegando / em porto / fundeado / retido / inativo / sem informacao]
  Rota      [Porto de origem] -> [Porto de destino]  |  ETA [data hora, se confirmado nos emails]
  Carga     [tipo e quantidade]

  URGENTE                          <- so se houver algo critico confirmado
  - [item critico curto e claro]

  OPERACIONAL
  - [apenas o confirmado nos emails: posicao, ROB, status da operacao, PDA]
  - Agente: [nome]  |  Contato: [email ou telefone, se disponivel]
  - Fretador: [nome, confirmado no fixture recap]
  - maximo 4 bullets
  - se nao houver informacao nova confirmada, escreva "sem atualizacao confirmada"

  PROXIMA VIAGEM
  - Se fixture confirmado:
    Tipo: [Voyage Charter / Time Charter]  |  Cliente: [se disponivel]
    Carga: [tipo e quantidade, se voyage charter]
    [Porto de carga] -> [Porto de descarga]
  - Se em negociacao (sem fixture assinado): 🔄 Em negociacao — [breve descricao se disponivel]
  - Se nao houver nenhuma informacao sobre proxima viagem: Em aberto

  PENDENCIAS
  - [apenas acoes pontuais pendentes que ainda nao foram realizadas, confirmadas nos emails]
  (omitir se nao houver nada pendente)

──────────────────────────────────────────────

[NIVEL]:
  🔴  retencao, danos, disputa ativa, prazo critico
  🟡  pendencia financeira, ETB incerto, decisao pendente
  🟢  operacao normal sem pendencias
  ⚫  sem viagem ativa ou sem informacao

ORDENACAO: 🔴 primeiro, 🟢 por ultimo.
Se nao houver URGENTE, omita essa secao inteira.

Apos o bloco de todos os navios, adicione (somente se houver itens):

──────────────────────────────────────────────
PENDENCIAS DE VIAGENS ANTERIORES
──────────────────────────────────────────────
[Navio] — VOY [XXX/XX] — [Contraparte]
  - [descricao do item: laytime, disputa, fatura, demurrage, despesas do armador, etc.]
  - Status: [situacao atual confirmada no email]

Ordenar por importancia:
1. Disputas juridicas ativas (advogados envolvidos, acordos formais pendentes)
2. Laytimes em aberto sem resposta dos afretadores
3. Notas de debito ou faturas aguardando resposta interna
4. Fechamentos de viagem com despesas pendentes
5. Outros itens administrativos

(omitir a secao inteira se nao houver pendencias de viagens anteriores)

---
Emails das ultimas 24 horas:
{email_content}"""

    response = client.messages.create(
        model="claude-sonnet-4-6",
        max_tokens=10000,
        messages=[{"role": "user", "content": prompt}],
    )
    return response.content[0].text


def send_report(report_text):
    today = datetime.now().strftime("%d/%m/%Y")
    msg = MIMEMultipart("alternative")
    msg["Subject"] = f"Relatorio Diario de Navios — {today}"
    msg["From"] = GMAIL_USER
    msg["To"] = REPORT_RECIPIENT
    msg["Cc"] = CRISTIANO_EMAIL

    html = f"<html><body><pre style='font-family:monospace;font-size:14px'>{report_text}</pre></body></html>"
    msg.attach(MIMEText(report_text, "plain", "utf-8"))
    msg.attach(MIMEText(html, "html", "utf-8"))

    recipients = list({REPORT_RECIPIENT, CRISTIANO_EMAIL})
    with smtplib.SMTP_SSL("smtp.gmail.com", 465) as server:
        server.login(GMAIL_USER, GMAIL_APP_PASSWORD)
        server.sendmail(GMAIL_USER, recipients, msg.as_string())

    print(f"[OK] Relatorio enviado para {', '.join(recipients)}")


def check_position_list_today(mail_conn):
    """Verifica se chegou position list do Cristiano ou Jerome hoje."""
    today = datetime.now().strftime("%d-%b-%Y")
    all_ids = []
    for sender in POSITION_LIST_SENDERS:
        _, ids = mail_conn.search(None, f'(FROM "{sender}" SINCE {today})')
        if ids[0]:
            all_ids.extend(ids[0].split())
    for msg_id in all_ids:
        _, msg_data = mail_conn.fetch(msg_id, "(RFC822)")
        msg = email.message_from_bytes(msg_data[0][1])
        for part in msg.walk():
            fn = part.get_filename()
            if fn:
                ext = decode_header_value(fn).lower().split(".")[-1]
                if ext in ("xlsx", "xls"):
                    return True
    return False


def check_report_already_sent_today(mail_conn):
    """Verifica se o relatorio diario ja foi enviado hoje."""
    today = datetime.now().strftime("%d-%b-%Y")
    sent_folders = ['"[Gmail]/Sent Mail"', '"[Gmail]/Enviados"', "Sent", "Enviados"]
    for folder in sent_folders:
        try:
            status, _ = mail_conn.select(folder)
            if status != "OK":
                continue
            _, ids = mail_conn.search(None, f'(SUBJECT "Relatorio Diario" SINCE {today})')
            mail_conn.select("inbox")
            if ids[0]:
                return True
        except Exception:
            try:
                mail_conn.select("inbox")
            except Exception:
                pass
    return False


if __name__ == "__main__":
    print(f"[{datetime.now().strftime('%d/%m/%Y %H:%M')}] Verificando position list...")

    if not GMAIL_APP_PASSWORD:
        print("[ERRO] GMAIL_APP_PASSWORD nao configurado.")
        exit(1)
    if not ANTHROPIC_API_KEY:
        print("[ERRO] ANTHROPIC_API_KEY nao configurado.")
        exit(1)

    mail = imaplib.IMAP4_SSL("imap.gmail.com")
    mail.login(GMAIL_USER, GMAIL_APP_PASSWORD)
    mail.select("inbox")

    if not check_position_list_today(mail):
        print("[INFO] Position list (Cristiano/Jerome) ainda nao chegou hoje. Aguardando.")
        mail.logout()
        exit(0)

    if check_report_already_sent_today(mail):
        print("[INFO] Relatorio ja enviado hoje. Nada a fazer.")
        mail.logout()
        exit(0)

    mail.logout()

    print("[INFO] Position list recebida. Gerando relatorio...")
    emails, positions_text, positions_meta, fleet_from_excel, last_report, last_report_date, fixture_recaps = fetch_vessel_emails()
    print(f"[INFO] {len(emails)} email(s) encontrado(s) nas ultimas 24h.")
    if fleet_from_excel:
        print(f"[INFO] Frota extraida da planilha: {', '.join(fleet_from_excel)}")
    else:
        print(f"[INFO] Frota padrao (fallback): {', '.join(FLEET)}")
    if positions_text:
        print(f"[INFO] Planilha de posicoes encontrada: {positions_meta}")
    else:
        print("[AVISO] Planilha de posicoes do Cristiano nao encontrada.")
    if last_report:
        print(f"[INFO] Ultimo relatorio encontrado: {last_report_date}")
    else:
        print("[AVISO] Ultimo relatorio nao encontrado.")
    print(f"[INFO] Fixture recaps encontrados: {len(fixture_recaps)}")

    report = generate_report(emails, positions_text, positions_meta, fleet_from_excel, last_report, last_report_date, fixture_recaps)
    print("\n--- RELATORIO GERADO ---")
    print(report.encode("cp1252", errors="replace").decode("cp1252"))
    print("------------------------\n")

    send_report(report)
